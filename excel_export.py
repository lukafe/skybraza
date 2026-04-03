"""
Geração de ficheiro Excel (.xlsx) para o pacote de exportação do escopo CertiK/IN 701.

Folhas:
  1. Resumo (Summary)       — instituição, trilha, data, KPIs
  2. No Escopo (In Scope)   — incisos sujeitos à auditoria com narrativas
  3. Fora do Escopo (Out)   — incisos excluídos e justificativa
  4. Prontidão Corpus       — status dos ficheiros de evidência por inciso

Uso:
    buf = build_scope_excel(scope_response, lang="pt")  # BytesIO
"""

from __future__ import annotations

import io
import unicodedata
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ── Paleta CertiK ────────────────────────────────────────────────────────────
_BLUE_DARK  = "0D2137"   # cabeçalhos principais
_BLUE_MID   = "1B4F72"   # cabeçalhos secundários
_BLUE_LIGHT = "D6E4F0"   # linhas alternadas
_ORANGE     = "E67E22"   # badges obrigatório
_GREEN      = "27AE60"   # badges acionado
_RED_LIGHT  = "FADBD8"   # fora de escopo
_WHITE      = "FFFFFF"
_GRAY_LIGHT = "F2F3F4"
_DARK_TEXT  = "1A1A1A"

_MANDATORY_FILL  = PatternFill("solid", fgColor="FAD7A0")
_TRIGGERED_FILL  = PatternFill("solid", fgColor="D5F5E3")
_OUT_FILL        = PatternFill("solid", fgColor=_RED_LIGHT)
_HEADER_FILL     = PatternFill("solid", fgColor=_BLUE_DARK)
_SUBHEADER_FILL  = PatternFill("solid", fgColor=_BLUE_MID)
_ALT_FILL        = PatternFill("solid", fgColor=_BLUE_LIGHT)
_TITLE_FONT      = Font(name="Calibri", size=14, bold=True, color=_WHITE)
_HEADER_FONT     = Font(name="Calibri", size=10, bold=True, color=_WHITE)
_SUBHEADER_FONT  = Font(name="Calibri", size=10, bold=True, color=_WHITE)
_BODY_FONT       = Font(name="Calibri", size=10, color=_DARK_TEXT)
_BOLD_FONT       = Font(name="Calibri", size=10, bold=True, color=_DARK_TEXT)
_MONO_FONT       = Font(name="Courier New", size=9, color=_DARK_TEXT)

_THIN = Side(style="thin", color="BDBDBD")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center")
_TOP_LEFT = Alignment(horizontal="left", vertical="top")


def _excel_formula_safe(s: str) -> str:
    """Evita que o Excel interprete células como fórmulas (=, +, -, @, tab)."""
    t = s.lstrip()
    if not t:
        return s
    if t[0] in "=\t\r" or (len(t) > 1 and t[0] in "+-" and t[1].isdigit()) or t.startswith("@"):
        return "'" + s
    return s


def _sanitize_cell_value(value: Any) -> Any:
    """Remove caracteres de controle e garante que strings são UTF-8 válido."""
    if value is None:
        return ""
    if isinstance(value, str):
        s = "".join(c for c in value if unicodedata.category(c)[0] != "C")
        return _excel_formula_safe(s)
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)


def _cell(ws, row: int, col: int, value: Any, *, font=None, fill=None,
          align=None, border=None, number_format: str | None = None):
    c = ws.cell(row=row, column=col, value=_sanitize_cell_value(value))
    if font:    c.font   = font
    if fill:    c.fill   = fill
    if align:   c.alignment = align
    if border:  c.border = border
    if number_format: c.number_format = number_format
    return c


def _write_header_row(ws, row: int, cols: list[str], fill=None):
    fill = fill or _SUBHEADER_FILL
    for ci, label in enumerate(cols, start=1):
        _cell(ws, row, ci, label, font=_HEADER_FONT, fill=fill,
              align=_CENTER, border=_THIN_BORDER)


def _auto_width(ws, max_width: int = 70):
    for col_cells in ws.columns:
        width = 8
        for c in col_cells:
            if c.value:
                w = min(len(str(c.value)), max_width)
                if w > width:
                    width = w
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width + 2


# ── i18n strings ─────────────────────────────────────────────────────────────
_STRINGS: dict[str, dict[str, str]] = {
    "sheet_summary":    {"pt": "Resumo",               "en": "Summary"},
    "sheet_in_scope":   {"pt": "No Escopo",            "en": "In Scope"},
    "sheet_out_scope":  {"pt": "Fora do Escopo",       "en": "Out of Scope"},
    "sheet_corpus":     {"pt": "Prontidão Corpus",     "en": "Corpus Readiness"},
    "title_scope":      {"pt": "Delimitação de Escopo IN 701 — CertiK",
                         "en": "IN 701 Scope Delimitation — CertiK"},
    "institution":      {"pt": "Instituição",          "en": "Institution"},
    "track":            {"pt": "Trilha",               "en": "Track"},
    "generated":        {"pt": "Gerado em",            "en": "Generated at"},
    "total_in_scope":   {"pt": "Incisos no escopo",    "en": "Clauses in scope"},
    "mandatory_n":      {"pt": "Obrigatórios",         "en": "Mandatory"},
    "triggered_n":      {"pt": "Acionados",            "en": "Triggered"},
    "out_scope_n":      {"pt": "Fora do escopo",       "en": "Out of scope"},
    "col_inciso_id":    {"pt": "Inciso ID",            "en": "Clause ID"},
    "col_item":         {"pt": "Item IN 701",          "en": "IN 701 Item"},
    "col_article":      {"pt": "Artigo IN 701",        "en": "IN 701 Article"},
    "col_origin":       {"pt": "Origem",               "en": "Origin"},
    "col_why":          {"pt": "Por que entra no escopo",
                         "en": "Why it is in scope"},
    "col_bcb":          {"pt": "Orientação BCB",       "en": "BCB Guidance"},
    "col_triggers":     {"pt": "Gatilhos",             "en": "Triggers"},
    "col_why_out":      {"pt": "Por que fora do escopo",
                         "en": "Why out of scope"},
    "col_corpus_status":{"pt": "Status Corpus",        "en": "Corpus Status"},
    "col_files":        {"pt": "Ficheiros",            "en": "Files"},
    "col_stub":         {"pt": "Usa STUB?",            "en": "Uses STUB?"},
    "col_missing":      {"pt": "Ficheiros ausentes",   "en": "Missing files"},
    "readiness_idx":    {"pt": "Índice prontidão",     "en": "Readiness index"},
    "yes":              {"pt": "Sim",                  "en": "Yes"},
    "no":               {"pt": "Não",                  "en": "No"},
    "track_intermediaria": {"pt": "Fase intermediária",   "en": "Intermediary phase"},
    "track_custodiante":   {"pt": "Trilha custodiante",   "en": "Custodian track"},
    "track_corretora":     {"pt": "Trilha corretora",     "en": "Broker/Exchange track"},
}


def _s(key: str, lang: str = "pt") -> str:
    d = _STRINGS.get(key, {})
    return d.get(lang) or d.get("pt") or key


# ── Sheet 1: Resumo / Summary ─────────────────────────────────────────────────
def _build_summary_sheet(ws, data: dict[str, Any], lang: str):
    ws.title = _s("sheet_summary", lang)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    inst   = data.get("institution") or "—"
    track  = data.get("track") or "intermediaria"
    resumo = data.get("resumo") or {}
    na     = resumo.get("total_sujeitos_auditoria", 0)
    mand   = resumo.get("obrigatorios_matriz", 0)
    cond   = resumo.get("acionados_por_respostas", 0)
    nf     = resumo.get("total_fora_escopo_auditoria", 0)
    gen    = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    cr     = data.get("corpus_readiness") or {}
    ri     = cr.get("readiness_index_0_100", "—")

    track_label = _s(f"track_{track}", lang)

    # Linha de título
    ws.merge_cells("A1:B1")
    _cell(ws, 1, 1, _s("title_scope", lang),
          font=_TITLE_FONT, fill=_HEADER_FILL, align=_CENTER)
    ws.row_dimensions[1].height = 28

    rows = [
        (_s("institution", lang), inst),
        (_s("track", lang),       track_label),
        (_s("generated", lang),   gen),
        (None, None),
        (_s("total_in_scope", lang),  na),
        (_s("mandatory_n", lang),     mand),
        (_s("triggered_n", lang),     cond),
        (_s("out_scope_n", lang),     nf),
        (_s("readiness_idx", lang),   f"{ri}%" if isinstance(ri, (int, float)) else ri),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        if k is None:
            continue
        c_key = ws.cell(row=i, column=1, value=k)
        c_key.font = _BOLD_FONT
        c_key.fill = _ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor=_WHITE)
        c_key.border = _THIN_BORDER
        c_key.alignment = _TOP_LEFT

        c_val = ws.cell(row=i, column=2, value=v)
        c_val.font = _BODY_FONT
        c_val.fill = _ALT_FILL if i % 2 == 0 else PatternFill("solid", fgColor=_WHITE)
        c_val.border = _THIN_BORDER
        c_val.alignment = _TOP_LEFT

    ws.freeze_panes = "A2"


# ── Sheet 2: No Escopo / In Scope ─────────────────────────────────────────────
def _build_in_scope_sheet(ws, incisos: list[dict[str, Any]], lang: str):
    ws.title = _s("sheet_in_scope", lang)

    headers = [
        _s("col_inciso_id", lang),
        _s("col_item", lang),
        _s("col_article", lang),
        _s("col_origin", lang),
        _s("col_why", lang),
        _s("col_bcb", lang),
        _s("col_triggers", lang),
    ]
    _write_header_row(ws, 1, headers, fill=_SUBHEADER_FILL)
    ws.row_dimensions[1].height = 20

    # Column widths
    widths = [12, 12, 18, 18, 60, 50, 20]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(incisos, start=2):
        origem = row.get("origem_escopo") or ""
        is_mandatory = "Obrigatório" in origem or "Mandatory" in origem
        fill = _MANDATORY_FILL if is_mandatory else _TRIGGERED_FILL
        alt_fill = PatternFill("solid", fgColor="FEF9E7") if is_mandatory else PatternFill("solid", fgColor="EAFAF1")

        vals = [
            row.get("inciso_id") or "",
            row.get("item") or "",
            row.get("artigo_in701") or "",
            origem,
            row.get("por_que_sera_auditado") or "",
            row.get("orientacao_relatorio_bcb") or "",
            ", ".join(row.get("perguntas_gatilho") or []),
        ]
        for ci, val in enumerate(vals, start=1):
            cell_fill = fill if ci == 1 else (alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor=_WHITE))
            fnt = _MONO_FONT if ci == 1 else _BODY_FONT
            aln = _CENTER if ci == 1 else _WRAP
            _cell(ws, ri, ci, val, font=fnt, fill=cell_fill, align=aln, border=_THIN_BORDER)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(len(incisos) + 1, 2)}"


# ── Sheet 3: Fora do Escopo / Out of Scope ────────────────────────────────────
def _build_out_scope_sheet(ws, incisos: list[dict[str, Any]], lang: str):
    ws.title = _s("sheet_out_scope", lang)

    headers = [
        _s("col_inciso_id", lang),
        _s("col_item", lang),
        _s("col_article", lang),
        _s("col_why_out", lang),
    ]
    _write_header_row(ws, 1, headers, fill=PatternFill("solid", fgColor="922B21"))
    ws.row_dimensions[1].height = 20

    widths = [12, 12, 18, 80]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, row in enumerate(incisos, start=2):
        alt = ri % 2 == 0
        vals = [
            row.get("inciso_id") or "",
            row.get("item") or "",
            row.get("artigo_in701") or "",
            row.get("por_que_nao_neste_escopo") or "",
        ]
        for ci, val in enumerate(vals, start=1):
            fill = _OUT_FILL if alt else PatternFill("solid", fgColor=_WHITE)
            fnt = _MONO_FONT if ci == 1 else _BODY_FONT
            aln = _CENTER if ci == 1 else _WRAP
            _cell(ws, ri, ci, val, font=fnt, fill=fill, align=aln, border=_THIN_BORDER)

    ws.freeze_panes = "A2"
    if incisos:
        ws.auto_filter.ref = f"A1:D{len(incisos) + 1}"


# ── Sheet 4: Prontidão Corpus / Corpus Readiness ──────────────────────────────
def _build_corpus_sheet(ws, corpus_readiness: dict[str, Any], lang: str):
    ws.title = _s("sheet_corpus", lang)

    headers = [
        _s("col_inciso_id", lang),
        _s("col_item", lang),
        _s("col_corpus_status", lang),
        _s("col_stub", lang),
        _s("col_files", lang),
        _s("col_missing", lang),
    ]
    _write_header_row(ws, 1, headers, fill=PatternFill("solid", fgColor="117A65"))
    ws.row_dimensions[1].height = 20

    widths = [12, 12, 16, 12, 55, 40]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    items = corpus_readiness.get("items") or []
    yes_s = _s("yes", lang)
    no_s  = _s("no", lang)

    for ri, row in enumerate(items, start=2):
        alt = ri % 2 == 0
        stub = row.get("uses_stub_reference", False)
        missing = row.get("ficheiros_ausentes_em_disco") or []
        vals = [
            row.get("inciso_id") or "",
            row.get("item") or "",
            row.get("corpus_status") or "",
            yes_s if stub else no_s,
            row.get("ficheiros_corpus") or "",
            "; ".join(missing) if missing else "",
        ]
        for ci, val in enumerate(vals, start=1):
            alt_fill = PatternFill("solid", fgColor="D1F2EB") if alt else PatternFill("solid", fgColor=_WHITE)
            bad_fill = PatternFill("solid", fgColor="FDEDEC")
            use_fill = bad_fill if (ci == 6 and missing) else alt_fill
            fnt = _MONO_FONT if ci == 1 else _BODY_FONT
            aln = _CENTER if ci in (1, 3, 4) else _WRAP
            _cell(ws, ri, ci, val, font=fnt, fill=use_fill, align=aln, border=_THIN_BORDER)

    ws.freeze_panes = "A2"
    if items:
        ws.auto_filter.ref = f"A1:F{len(items) + 1}"


# ── Entry point ───────────────────────────────────────────────────────────────
def build_scope_excel(scope_response: dict[str, Any], lang: str = "pt") -> io.BytesIO:
    """
    Recebe a resposta completa do ``POST /scope`` e devolve um BytesIO com o ficheiro .xlsx.
    """
    wb = Workbook()

    ws_sum = wb.active
    _build_summary_sheet(ws_sum, scope_response, lang)

    ws_in  = wb.create_sheet()
    _build_in_scope_sheet(ws_in,  scope_response.get("incisos_sujeitos_auditoria") or [], lang)

    ws_out = wb.create_sheet()
    _build_out_scope_sheet(ws_out, scope_response.get("incisos_fora_escopo_auditoria") or [], lang)

    ws_corp = wb.create_sheet()
    _build_corpus_sheet(ws_corp, scope_response.get("corpus_readiness") or {}, lang)

    for ws in wb.worksheets:
        _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


__all__ = ["build_scope_excel"]
